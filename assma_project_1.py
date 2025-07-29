import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook

#العملات المتاحة 
currencies = ['USD', 'EUR', 'KWD', 'IQD', 'SYP', 'CAD']

# أسعار الصرف الثابتة مقابل الدولار الأمريكي (USD)
usd_rates = {
    'USD': 1,
    'EUR': 0.9,
    'KWD': 0.30,
    'IQD': 1390,
    'SYP': 1000,
    'CAD': 1.35,
}

# دالة لحساب سعر الصرف بين عملتين عن طريق تحويلهم أولاً إلى USD
def get_rate_manual(from_curr, to_curr):
    if from_curr not in usd_rates or to_curr not in usd_rates:
        raise ValueError("هذه العملة غير مدعومة حالياً.")

    rate_from = usd_rates[from_curr]
    rate_to = usd_rates[to_curr]

    return rate_to / rate_from

conversion_history = []

def save_record_to_excel(timestamp, conversion_text):
    filename = "سجل_التحويلات.xlsx"

    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "سجل التحويلات"
        ws.append(["التاريخ والوقت", "التحويل"])

    ws.append([timestamp, conversion_text])
    wb.save(filename)

def save_to_excel():
    if not conversion_history:
        messagebox.showinfo("لا يوجد بيانات", "سجل التحويلات فارغ.")
        return

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "سجل التحويلات"
        ws.append(["التاريخ والوقت", "التحويل"])

        for record in conversion_history:
            if " | " in record:
                time_part, convert_part = record.split(" | ", 1)
                ws.append([time_part, convert_part])
            else:
                ws.append(["", record])

        filename = "سجل_التحويلات.xlsx"
        wb.save(filename)
        messagebox.showinfo("تم الحفظ", f"تم حفظ الملف بنجاح باسم:\n{filename}")
    except Exception as e:
        messagebox.showerror("خطأ", f"فشل حفظ الملف: {str(e)}")

def show_history():
    if not conversion_history:
        messagebox.showinfo("سجل التحويلات", "لا يوجد تحويلات حتى الآن.")
        return

    history_window = tk.Toplevel(root)
    history_window.title("سجل التحويلات")
    history_window.geometry("400x300")
    history_window.resizable(False, False)

    tk.Label(history_window, text="سجل التحويلات:", font=("Arial", 12, "bold")).pack(pady=10)

    text_area = tk.Text(history_window, wrap="word")
    text_area.pack(expand=True, fill='both', padx=10, pady=10)

    for entry in conversion_history:
        text_area.insert(tk.END, entry + "\n")
    text_area.config(state='disabled')

def convert_currency():
    try:
        amount = float(entry_amount.get())
        from_curr = combo_from.get()
        to_curr = combo_to.get()

        if from_curr == "" or to_curr == "":
            messagebox.showwarning("تنبيه", "يرجى اختيار العملة.")
            return

        rate = get_rate_manual(from_curr, to_curr)
        result = amount * rate
        result_rounded = round(result, 2)

        label_result.config(text=f"{amount} {from_curr} = {result_rounded} {to_curr}")

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        record = f"{now} | {amount} {from_curr} -> {result_rounded} {to_curr}"
        conversion_history.append(record)

        save_record_to_excel(now, f"{amount} {from_curr} -> {result_rounded} {to_curr}")

        label_rate.config(text=f"سعر الصرف: 1 {from_curr} = {round(rate, 4)} {to_curr}")

    except ValueError as ve:
        messagebox.showerror("خطأ", str(ve))
    except Exception as e:
        messagebox.showerror("خطأ", f"حدث خطأ: {str(e)}")

def update_exchange_rate(*args):
    from_curr = combo_from.get()
    to_curr = combo_to.get()

    if from_curr and to_curr:
        try:
            rate = get_rate_manual(from_curr, to_curr)
            label_rate.config(text=f"سعر الصرف: 1 {from_curr} = {round(rate, 4)} {to_curr}")
        except Exception:
            label_rate.config(text="سعر الصرف غير متاح حالياً.")

root = tk.Tk()
root.title("برنامج صرافة")
root.geometry("400x470")
root.resizable(False, False)

tk.Label(root, text="محول العملات", font=("Arial", 16, "bold")).pack(pady=10)

tk.Label(root, text="المبلغ:").pack()
entry_amount = tk.Entry(root, justify='center')
entry_amount.pack(pady=5)

tk.Label(root, text="من:").pack()
combo_from = ttk.Combobox(root, values=currencies, state='readonly', justify='center')
combo_from.pack(pady=5)

tk.Label(root, text="إلى:").pack()
combo_to = ttk.Combobox(root, values=currencies, state='readonly', justify='center')
combo_to.pack(pady=5)

label_rate = tk.Label(root, text="", fg="blue")
label_rate.pack(pady=5)

tk.Button(root, text="تحويل", command=convert_currency, bg='green', fg='white').pack(pady=10)

label_result = tk.Label(root, text="", font=("Arial", 12, "bold"))
label_result.pack(pady=10)

tk.Button(root, text="عرض سجل التحويلات", command=show_history).pack(pady=5)

tk.Button(root, text="حفظ السجل إلى Excel", command=save_to_excel, bg="blue", fg="white").pack(pady=5)

combo_from.bind("<<ComboboxSelected>>", update_exchange_rate)
combo_to.bind("<<ComboboxSelected>>", update_exchange_rate)

root.mainloop()